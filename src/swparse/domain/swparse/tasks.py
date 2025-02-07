from __future__ import annotations

import os
import io
import re
import json
import base64
import tempfile
from uuid import uuid4
from typing import TYPE_CHECKING, Optional
from pydantic import TypeAdapter, ValidationError


import pymupdf
import html_text
import mammoth
import mistletoe
import structlog
import pandas as pd
import pypdfium2 as pdfium
from unoserver import client
from html2text import html2text
from PIL import Image as PILImage
from openpyxl import load_workbook
import markdown as markdown_converter
from markdownify import markdownify as md


from swparse.config.app import settings
from swparse.db.models import ContentType
from swparse.domain.swparse.convert import pdf_markdown
from swparse.domain.swparse.utils import (
    change_file_ext,
    convert_pptx_to_md,
    convert_xls_to_xlsx_bytes,
    extract_tables_gliner,
    get_file_content,
    get_file_name,
    extract_excel_images,
    save_file,
    read_file,
    save_metadata,
    get_metadata
)
from  swparse.domain.swparse.schemas import Page, LLAMAJSONOutput

if TYPE_CHECKING:
    from saq.types import Context

logger = structlog.get_logger()
BUCKET = settings.storage.BUCKET
MINIO_ROOT_USER = settings.storage.ROOT_USER
MINIO_ROOT_PASSWORD = settings.storage.ROOT_PASSWORD
MEMORY_USAGE_LOG  = settings.app.MEMORY_USAGE_LOG
SAQ_PROCESSES = settings.saq.PROCESSES


async def parse_xlsx_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None, sheet_index: Optional[list[str|int]]= None) -> dict[str, str]:

    logger.info("Started parse_xlsx_s3")
    results = {}
    try:
        content =  await read_file(s3_url)

        if isinstance(content, str):
            content = content.encode()

        if ext == "application/vnd.ms-excel":
            content = await convert_xls_to_xlsx_bytes(content)

        file_name = await get_file_name(s3_url)
        str_buffer = io.BytesIO(content)

        if sheet_index is None:

            pxl_doc = load_workbook(filename=str_buffer)
            sheet_index = pxl_doc.sheetnames
            sheet_index = [] if sheet_index is None else sheet_index

        csv_content = ""
        html_content = ""
        md_content = ""
        all_images = {}
        for sheet_name in sheet_index:
            try:
                df = pd.read_excel(str_buffer, sheet_name=sheet_name, header=0, na_filter=False)
            except Exception as e:

                if sheet_name.isdigit():
                    sheet_name = int(sheet_name)
                    try:
                        df = pd.read_excel(str_buffer, sheet_name=sheet_name, header=0, na_filter=False)
                    except ValueError as e:
                        logger.error(f"Sheet index {sheet_name} is out of range.")
                        continue
                else:
                    logger.error(f"sheet: {sheet_name} is not found in the provided file!")
                    continue

            images = await extract_excel_images(str_buffer, sheet_name)
   
            all_images.update(images)
            csv_content += df.to_csv(index=False, na_rep="")
            csv_content += "\n"

            md_content += df.to_markdown(index=False)
            md_content += "\n"

            html_content += df.to_html(index=False, na_rep="")
            html_content += "<br>"

            # Adding extracted images
            for image_key in images.keys():
                md_content += f"\n![{image_key}]({image_key})\n\n"

                html_content += f'<br><img src=\"{image_key}" alt=\"{image_key}\" /><br><br>'

        csv_file_name = await change_file_ext(file_name, "csv")
        csv_file_path = await save_file(csv_file_name, csv_content)

        html_file_name = await change_file_ext(file_name, "html")
        html_file_path = await save_file(html_file_name, html_content)

        md_file_name = await change_file_ext(file_name, "md")
        md_file_path = await save_file(md_file_name, md_content)

        # Parsing to Text
        text_content = html_text.extract_text(html_content)
        text_file_name = await change_file_ext(file_name, "txt")
        txt_file_path = await save_file(text_file_name, text_content)

        results = {
            ContentType.CSV.value: csv_file_path,
            ContentType.MARKDOWN.value: md_file_path,
            ContentType.HTML.value: html_file_path,
            ContentType.TEXT.value: txt_file_path
        }

        if all_images:
            img_file_name = await change_file_ext(file_name, "json")
            img_file_path = await save_file(img_file_name, json.dumps(all_images))
            results[ContentType.IMAGES.value] = img_file_path

        if table_query:
            tables_content = extract_tables_gliner(table_query["tables"], md_content, table_query["output"])
            tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
            tables_file_path = await save_file(tables_file_name, tables_content)
            results[table_query["raw"]] = tables_file_path

        await save_metadata(s3_url, metadata=results)

    except Exception as e:
        logger.error(f"Error parsing XLSX file from S3: {e}")
        raise

    logger.info("Extracted results")
    logger.info(results)
    return results


async def extract_string(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]:


    logger.info("Started extract_string")
    file_name = await get_file_name(s3_url)
    txt_file_name = await change_file_ext(file_name, "txt")

    byte_string = await read_file(s3_url)
       
    try:
        out_txt = str(byte_string.decode())
        text_file_path = await save_file(txt_file_name, out_txt)
    except UnicodeDecodeError:
        out_txt = str(byte_string)
        text_file_path = await save_file(txt_file_name, out_txt)
    results = {ContentType.TEXT.value: text_file_path}

    await save_metadata(s3_url, metadata=results)

    return results


async def _pdf_exchange(s3_url: str, start_page: int = 0, end_page: int = 40, force_ocr:bool = False) -> dict[str, str]:

    file_name = await get_file_name(s3_url)
    logger.info("Reading content with thread")
    content =  await read_file(s3_url)

    logger.info("Finished reading content")

    results:LLAMAJSONOutput  = await pdf_markdown(content, start_page=start_page, max_pages=end_page, ocr_all_pages=force_ocr)
    logger.info("Finished parsing")
    data:dict[str, str] = {}

    md_file_name = await change_file_ext(file_name, "md")

    data[ContentType.MARKDOWN.value] = await save_file(md_file_name, results.markdown)
    # HTML results
    html_file_name = await change_file_ext(file_name, "html")
    data[ContentType.HTML.value] = await save_file(html_file_name, results.html)
    # Text Parsing
    txt_file_name = await change_file_ext(file_name, "txt")
    data[ContentType.TEXT.value] = await save_file(txt_file_name, results.text)

    try:
        page_adapter = TypeAdapter(list[Page])
        page_adapter.validate_python(results.pages)
        json_file_name = await change_file_ext(file_name, "json")
        data[ContentType.JSON.value]  = await save_file(json_file_name, json.dumps(results.pages))

        logger.info("Validation successful!\n")

    except ValidationError as e:
        logger.error("Validation failed!")
        logger.error(e.json())

    # Saving image metadata
    img_file_name = await change_file_ext(file_name, "json")
    data[ContentType.IMAGES.value] = await save_file(img_file_name, json.dumps(results.images))

    return data

 
async def parse_docx_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]: 

 
    logger.info("Started parse_docx_s3")
    file_name = await get_file_name(s3_url)

    # HTML parsing
    byte_content = await read_file(s3_url)
    with io.BytesIO(byte_content) as byte_stream:
        html_data = mammoth.convert_to_html(byte_stream)  # type: ignore
 
    htmlData: str = html_data.value  # type: ignore
    # TODO: refactor using a html tree

    img_tags = re.findall(r'<img\s+[^>]*src=[\'"].+?[\'"][^>]*>', htmlData)

    images = {}

    for i, img in enumerate(img_tags):

        src = re.search(r'src=[\'"]([^\'"]+)[\'"]', img).group(1)
        header, encoded = src.split(",", 1)
        image_type = header.split(";")[0].split(":")[1].split("/")[1]
        image_bytes = base64.b64decode(encoded)

        image_key = f"image-{i}.{image_type}"
        htmlData = htmlData.replace(img, f'<img src="{image_key}" alt="{image_key}" />', 1)
        image_file_path = await save_file(image_key, image_bytes)
        images[image_key] = image_file_path


    html_file_name = await change_file_ext(file_name, "html")
    html_file_path = await save_file(html_file_name, htmlData)

    # Markdown parsing
    markdown = md(htmlData)  # type: ignore
    md_file_name = await change_file_ext(file_name, "md")
    md_file_path = await save_file(md_file_name, markdown)

    # Parsing to Text
    text_content = html_text.extract_text(htmlData, guess_layout=True)
    text_file_name = await change_file_ext(file_name, "txt")
    txt_file_path = await save_file(text_file_name, text_content)

    # Saving image metadata
    file_name = file_name + "_image_metadata"
    img_file_name = await change_file_ext(file_name, "json")
    img_file_path = await save_file(img_file_name, json.dumps(images))


    results = {
        ContentType.HTML.value: html_file_path,
        ContentType.MARKDOWN.value: md_file_path,
        ContentType.TEXT.value: txt_file_path,
        ContentType.IMAGES.value: img_file_path,
    }

    if table_query:
        tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
        tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
        tables_file_path = await save_file(tables_file_name, tables_content)
        results[table_query["raw"]] = tables_file_path
 
    await save_metadata(s3_url, metadata=results)
    logger.info("Extracted results")
    logger.info(results)
    return results


async def parse_pdf_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None, force_ocr: bool = False, plain_text: bool = False) -> dict[str, str]:
    logger.info("Started parse_pdf_s3")

    file_name = await get_file_name(s3_url)
    if plain_text:
        content = await read_file(s3_url)

        with pymupdf.open(stream=io.BytesIO(content), filetype="pdf") as doc:
            text = chr(12).join([page.get_text("text") for page in doc])

        text_file_name = await change_file_ext(file_name, "txt")
        txt_file_path = await save_file(text_file_name, text)

        return { ContentType.TEXT.value: txt_file_path}

    results = await _pdf_exchange(s3_url, force_ocr= force_ocr)
  
    if table_query:
        markdown = await get_file_content(results["markdown"])
        tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
        tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
        tables_file_path = await save_file(tables_file_name, tables_content)
        results[table_query["raw"]] = tables_file_path
 
    await save_metadata(s3_url, metadata=results)
    logger.info("Extracted results")
    logger.info(results)
    return results


async def parse_pdf_page_s3(ctx: Context, *, s3_url: str, page: int) -> dict[str, str]:
    return await _pdf_exchange(s3_url, start_page=page)


async def parse_image_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None, force_ocr:bool = False) -> dict[str, str]:

    logger.info("Started parse_image_s3")
    content_byte = await read_file(s3_url)

    pil_image = PILImage.open(io.BytesIO(content_byte)).convert("RGB")
    pdf = pdfium.PdfDocument.new()

    image = pdfium.PdfImage.new(pdf)
    image.set_bitmap(pdfium.PdfBitmap.from_pil(pil_image))
    width, height = image.get_size()

    matrix = pdfium.PdfMatrix().scale(width, height)
    image.set_matrix(matrix)

    page = pdf.new_page(width, height)
    page.insert_obj(image)
    page.gen_content()
    
    pdf_buffer = io.BytesIO()
    pdf.save(pdf_buffer)
    pdf_buffer.seek(0) 
    
    pdf_s3_url = await change_file_ext(s3_url, "pdf")
    pdf_content = pdf_buffer.read()
    pdf_s3_url = await save_file(pdf_s3_url, pdf_content, randomize=False)
    results = await _pdf_exchange(pdf_s3_url, force_ocr)

    if table_query:
        file_name = await get_file_name(s3_url)
        markdown = await get_file_content(results["markdown"])
        tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
        tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
        tables_file_path = await save_file(tables_file_name, tables_content)
        results[table_query["raw"]] = tables_file_path

    logger.info("Extracted results")
    logger.info(results)
    await save_metadata(s3_url, results)
    
    return results


async def extract_text_files(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]:
   
    logger.info("Started extract_text_files")
    results = {}
    try:
    
        file_name = await get_file_name(s3_url)
       
        content = await read_file(s3_url)
        if isinstance(content, bytes):
            content = content.decode("utf-8")

        text_file_name = await change_file_ext(file_name, "txt")
        text_file_path = await save_file(text_file_name, content)

        if ext == "text/xml":
            df = pd.read_xml(io.StringIO(content))
            html_content = df.to_html(index=False)

        elif ext == "text/csv":
            csv_buffer = io.StringIO(content)
            df = pd.read_csv(csv_buffer, index_col=False)
            txt_content = df.to_string(index=False)
            text_file_name = await change_file_ext(file_name, "txt")
            text_file_path = await save_file(text_file_name, txt_content)
            html_content = df.to_html(index=False)

        else:
            html_content = markdown_converter.markdown(content)
        html_file_name = await change_file_ext(file_name, "html")
        html_file_path = await save_file(html_file_name, html_content)

        # Markdown Parsing
        markdown = html2text(html_content)
        md_file_name = await change_file_ext(file_name, "md")
        md_file_path = await save_file(md_file_name, markdown)

        results = {
            ContentType.MARKDOWN.value: md_file_path,
            ContentType.TEXT.value: text_file_path,
            ContentType.HTML.value: html_file_path,
        }
        if table_query:
            tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
            tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
            tables_file_path = await save_file(tables_file_name, tables_content)
            results[table_query["raw"]] = tables_file_path


        await save_metadata(s3_url, metadata=results)
        logger.info("Extracted results")
        logger.info(results)

    except Exception as e:
        logger.exception(f"Error while parsing document: {e}")

    return results


async def parse_doc_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]:
    logger.info("Started parse_doc_s3")

    file_name = await get_file_name(s3_url)
    content_byte = await read_file(s3_url)
 
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_input_path = os.path.join(temp_dir, file_name)
        
        with open(temp_input_path, "wb") as local_file:
            local_file.write(content_byte)
                
 
        conv = client.UnoClient(server="libreoffice", port="2003", host_location="remote")
        results = {}

        txt_name = await change_file_ext(file_name, "txt")
        temp_txt_path = os.path.join(temp_dir, txt_name)
        conv.convert(inpath=temp_input_path, outpath=temp_txt_path)
        with open(temp_txt_path, "rb") as converted_file:
            txt_s3_path = await save_file(txt_name, converted_file.read())
        results[ContentType.TEXT.value] = txt_s3_path

        html_name = await change_file_ext(file_name, "html")
        temp_html_path = os.path.join(temp_dir, html_name)
        conv.convert(inpath=temp_input_path, outpath=temp_html_path)
        with open(temp_html_path, "rb") as converted_file:
            html_s3_path = await save_file(html_name, converted_file.read())
        results[ContentType.HTML.value] = html_s3_path

        with open(temp_html_path) as html_file:
            markdown = md(html_file.read())
            md_file_name = await change_file_ext(file_name, "md")
            md_file_path = await save_file(md_file_name, markdown)
        results[ContentType.MARKDOWN.value] = md_file_path

        if table_query:
            tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
            tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
            tables_file_path = await save_file(tables_file_name, tables_content)
            results[table_query["raw"]] = tables_file_path

  
        await save_metadata(s3_url, metadata=results)
        logger.info("Extracted results")
        logger.info(results)
    return results


async def parse_ppt_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]:
    logger.info("Started parse_ppt_s3")
 
    file_name = await get_file_name(s3_url)
    byte_content = await read_file(s3_url)

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_input_path = os.path.join(temp_dir, file_name)
   
        with open(temp_input_path, "wb") as local_file:
            local_file.write(byte_content)

        conv = client.UnoClient(server="libreoffice", port="2003", host_location="remote")

        results = {}

        pptx_name = await change_file_ext(file_name, "pptx")
        temp_pptx_path = os.path.join(temp_dir, pptx_name)
        conv.convert(inpath=temp_input_path, outpath=temp_pptx_path)
        with open(temp_pptx_path, "rb") as converted_file:
            pptx_s3_path = await save_file(pptx_name, converted_file.read())
        results[ContentType.TEXT.value] = pptx_s3_path
        md_file_name = await change_file_ext(file_name, "md")
        new_uuid = uuid4()
        md_file_path = f"{BUCKET}/{new_uuid}_{md_file_name}"
        
        pptx_content = await read_file(pptx_s3_path)
        pptx_file_like = io.BytesIO(pptx_content)  
        markdown = convert_pptx_to_md(pptx_file_like, file_name)
        
        html_results = mistletoe.markdown(markdown)
        text_results = html_text.extract_text(html_results, guess_layout=True)

        html_file_name = await change_file_ext(file_name, "html")
        html_file_path = await save_file(html_file_name, html_results)

        txt_file_name = await change_file_ext(file_name, "txt")
        txt_file_path = await save_file(txt_file_name, text_results)

        results = {
            ContentType.MARKDOWN.value: md_file_path,
            ContentType.HTML.value: html_file_path,
            ContentType.TEXT.value: txt_file_path,
        }
     
        await save_metadata(s3_url, metadata=results)
        logger.info("Extracted results")
        logger.info(results)
        return results


async def parse_pptx_s3(ctx: Context, *, s3_url: str, ext: str, table_query: dict | None) -> dict[str, str]:
    logger.info("Started parse_pptx_s3")
    
    file_name = await get_file_name(s3_url)
    md_file_name = await change_file_ext(file_name, "md")
    
    pptx_content = await read_file(s3_url)
    pptx_file_like = io.BytesIO(pptx_content)  
    
    markdown_content = convert_pptx_to_md(pptx_file_like, file_name)
    md_file_path = await save_file(md_file_name, markdown_content)

    html_content = mistletoe.markdown(markdown_content)

    text_content = html_text.extract_text(html_content, guess_layout=True)

    html_file_name = await change_file_ext(file_name, "html")
    html_file_path = await save_file(html_file_name, html_content)

    txt_file_name = await change_file_ext(file_name, "txt")
    txt_file_path = await save_file(txt_file_name, text_content)

    results = {
        ContentType.MARKDOWN.value: md_file_path,
        ContentType.HTML.value: html_file_path,
        ContentType.TEXT.value: txt_file_path,
    }

    if table_query:
        tables_content = extract_tables_gliner(table_query["tables"], markdown_content, table_query["output"])
        tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
        tables_file_path = await save_file(tables_file_name, tables_content)
        results[table_query["raw"]] = tables_file_path
 
    await save_metadata(s3_url, metadata=results)
    logger.info("Extracted results")
    logger.info(results)
    return results


async def get_extracted_url(ctx: Context, *, s3_url: str, table_query: dict | None) -> dict[str, str]:
    logger.info("working get_extracted_url")

    metadata = await get_metadata(s3_url=s3_url)
  
    image_file_path = metadata.get("images")
    # images metadata normalization
    if image_file_path:
        try:
            image_metadata = json.loads(image_file_path)
            file_name = await get_file_name(s3_url)
            img_file_name = await change_file_ext(file_name, "json")
            img_file_path = await save_file(img_file_name, json.dumps(image_metadata))
            metadata["images"]=img_file_path
            await save_metadata(s3_url,  metadata)
        except Exception as err:
            # correct filepath
            pass


    if table_query:
        file_name = await get_file_name(s3_url)
        markdown = await get_file_content(metadata["markdown"])
        tables_content = extract_tables_gliner(table_query["tables"], markdown, table_query["output"])
        tables_file_name = await change_file_ext("extracted_tables_" + file_name, table_query["output"])
        tables_file_path = await save_file(tables_file_name, tables_content)
        metadata[table_query["raw"]] = tables_file_path

    return metadata


async def extract_advanced_tables(ctx: Context, *, markdown: str, table_query: dict) -> dict[str, str]:
    return extract_tables_gliner(table_query, markdown)
